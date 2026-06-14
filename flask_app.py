from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import pandas as pd
import io
import openpyxl                                          # ← NUEVO
from datetime import datetime

# ── ReportLab imports (necesarios para el PDF de prácticas) ──
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable
)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'practicas-secret-2026')

DATABASE_URL = os.environ.get('DATABASE_URL')

def get_db_connection():
    if not DATABASE_URL:
        raise ValueError("DATABASE_URL no está configurada en las variables de entorno")
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS estudiantes (
            id SERIAL PRIMARY KEY,
            cedula TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            sitio TEXT,
            programa TEXT DEFAULT 'Fisioterapia',
            sede TEXT,
            nivel_practica TEXT,
            grupo TEXT,
            correo TEXT,
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS docentes (
            id SERIAL PRIMARY KEY,
            documento TEXT UNIQUE,
            nombre TEXT NOT NULL,
            correo TEXT,
            estado TEXT DEFAULT 'Activo'
        );

        CREATE TABLE IF NOT EXISTS escenarios (
            id SERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            direccion TEXT,
            cupos INTEGER DEFAULT 10,
            estado TEXT DEFAULT 'Activo'
        );

        CREATE TABLE IF NOT EXISTS asignaciones (
            id SERIAL PRIMARY KEY,
            estudiante_id INTEGER REFERENCES estudiantes(id) ON DELETE CASCADE,
            docente_id INTEGER REFERENCES docentes(id) ON DELETE SET NULL,
            escenario_id INTEGER REFERENCES escenarios(id) ON DELETE SET NULL,
            nivel_practica TEXT,
            grupo TEXT,
            rotacion INTEGER NOT NULL,
            horario TEXT,
            fecha_inicio DATE,
            fecha_fin DATE,
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    conn.commit()
    cur.close()
    conn.close()

with app.app_context():
    init_db()


# ==============================================================
# HELPERS PARA EL PDF DE PRÁCTICAS (lee los Excel originales)
# ==============================================================

# Paleta de colores
_AZUL_HEADER   = colors.HexColor('#003366')
_AZUL_ROTACION = colors.HexColor('#1F4E79')
_VERDE_DOCENTE = colors.HexColor('#E2EFDA')
_GRIS_ALTERNO  = colors.HexColor('#F2F2F2')
_BLANCO        = colors.white
_AZUL_TITULO   = colors.HexColor('#2E75B6')

# Ruta a los Excel — ajusta si están en otra carpeta de tu proyecto
EXCEL_FILES = [
    'PROGRAMACIÓN_PRÁCTICA_I_2026-1.xlsx',
    'PROGRAMACIÓN_PRÁCTICA_II_2026-1.xlsx',
    'PROGRAMACIÓN_PRÁCTICA_III_2026-1.xlsx',
    'PROGRAMACIÓN_PRÁCTICA_INTEGRAL_2026-1.xlsx',
]
# Carpeta donde están los xlsx (al lado del app.py por defecto)
EXCEL_DIR = os.path.join(os.path.dirname(__file__), 'uploads')


def _parse_sheet(ws):
    """Lee un sheet de openpyxl y extrae header_lines y rotaciones."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([c if c is not None else '' for c in row])

    header_lines, rotaciones = [], []
    current_rot = None
    i = 0

    while i < len(rows):
        row = rows[i]
        first = str(row[0]).strip()

        if all(str(c).strip() == '' for c in row):
            i += 1
            continue

        if first.lower().startswith('rotación') or first.lower().startswith('rotacion'):
            if current_rot:
                rotaciones.append(current_rot)
            current_rot = {'titulo': first, 'nota': '', 'escenarios': []}
            i += 1
            continue

        if current_rot is None and first and not any(str(c).strip() for c in row[1:]):
            header_lines.append(first)
            i += 1
            continue

        if current_rot is not None and first and not any(str(c).strip() for c in row[1:]):
            current_rot['nota'] = first
            i += 1
            continue

        non_empty = [str(c).strip() for c in row if str(c).strip()]
        if len(non_empty) >= 2 and current_rot is not None:
            next_row = rows[i + 1] if i + 1 < len(rows) else []
            next_vals = [str(c).strip() for c in next_row if str(c).strip()]
            if any('docente' in v.lower() for v in next_vals):
                escenarios_nombres = [str(c).strip() for c in row]
                i += 1
                docentes_row = [str(c).strip() for c in rows[i]]
                i += 1

                escenarios = []
                for col_idx, nombre in enumerate(escenarios_nombres):
                    if nombre:
                        docente = docentes_row[col_idx] if col_idx < len(docentes_row) else ''
                        docente = docente.replace('Docente:', '').replace('Docente :', '').strip()
                        escenarios.append({'col': col_idx, 'nombre': nombre,
                                           'docente': docente, 'estudiantes': []})

                while i < len(rows):
                    r = rows[i]
                    r_first = str(r[0]).strip()
                    if all(str(c).strip() == '' for c in r):
                        i += 1
                        break
                    if r_first.lower().startswith('rotación') or r_first.lower().startswith('rotacion'):
                        break
                    if r_first and not any(str(c).strip() for c in r[1:]):
                        current_rot['nota'] = r_first
                        i += 1
                        break
                    for esc in escenarios:
                        col = esc['col']
                        val = str(r[col]).strip() if col < len(r) else ''
                        if val:
                            esc['estudiantes'].append(val)
                    i += 1

                current_rot['escenarios'] = escenarios
                continue

        if current_rot is None:
            texto = ' | '.join(str(c).strip() for c in row if str(c).strip())
            if texto:
                header_lines.append(texto)

        i += 1

    if current_rot:
        rotaciones.append(current_rot)

    return header_lines, rotaciones


def _build_rotation_table(rotacion, styles):
    """Construye los elementos ReportLab de una rotación."""
    elements = []
    escenarios = rotacion['escenarios']
    n_cols = len(escenarios)
    if n_cols == 0:
        return elements

    PAGE_W = landscape(letter)[0] - 2 * cm
    col_w = PAGE_W / n_cols

    rot_style = ParagraphStyle('rot', parent=styles['Normal'],
                               fontName='Helvetica-Bold', fontSize=9,
                               textColor=_BLANCO, alignment=TA_LEFT)
    esc_style = ParagraphStyle('esc', parent=styles['Normal'],
                               fontName='Helvetica-Bold', fontSize=8,
                               textColor=_BLANCO, alignment=TA_CENTER)
    doc_style = ParagraphStyle('doc', parent=styles['Normal'],
                               fontName='Helvetica-BoldOblique', fontSize=7.5,
                               textColor=colors.HexColor('#1F4E79'), alignment=TA_CENTER)
    est_style = ParagraphStyle('est', parent=styles['Normal'],
                               fontName='Helvetica', fontSize=7.5,
                               alignment=TA_CENTER)

    header_row  = [Paragraph(esc['nombre'], esc_style) for esc in escenarios]
    docente_row = [Paragraph(esc['docente'], doc_style) for esc in escenarios]

    max_est = max((len(esc['estudiantes']) for esc in escenarios), default=0)
    student_rows = []
    for row_i in range(max_est):
        student_rows.append([
            Paragraph(esc['estudiantes'][row_i] if row_i < len(esc['estudiantes']) else '', est_style)
            for esc in escenarios
        ])

    data = [header_row, docente_row] + student_rows
    ts = [
        ('BACKGROUND', (0, 0), (-1, 0), _AZUL_HEADER),
        ('TEXTCOLOR',  (0, 0), (-1, 0), _BLANCO),
        ('BACKGROUND', (0, 1), (-1, 1), _VERDE_DOCENTE),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFBFBF')),
        ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',   (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    for ri in range(2, len(data)):
        ts.append(('BACKGROUND', (0, ri), (-1, ri),
                   _GRIS_ALTERNO if ri % 2 == 0 else _BLANCO))

    table = Table(data, colWidths=[col_w] * n_cols)
    table.setStyle(TableStyle(ts))

    rot_title = Table([[Paragraph(rotacion['titulo'], rot_style)]], colWidths=[PAGE_W])
    rot_title.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), _AZUL_ROTACION),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 6),
    ]))

    elements.append(rot_title)
    elements.append(table)

    if rotacion.get('nota'):
        nota_style = ParagraphStyle('nota', parent=styles['Normal'],
                                    fontName='Helvetica-Oblique', fontSize=7.5,
                                    textColor=colors.HexColor('#7F7F7F'))
        elements.append(Spacer(1, 3))
        elements.append(Paragraph(f"⚠ {rotacion['nota']}", nota_style))

    elements.append(Spacer(1, 10))
    return elements


# ==================== DASHBOARD ====================
@app.route('/')
def index():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT a.id, e.nombre as estudiante, e.cedula, d.nombre as docente, 
               es.nombre as escenario, a.rotacion, a.horario, a.fecha_inicio, a.fecha_fin
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY a.fecha_creacion DESC LIMIT 10
    ''')
    asignaciones = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('index.html', asignaciones=asignaciones)


# ==================== ESTUDIANTES CRUD ====================
@app.route('/estudiantes')
def estudiantes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('estudiantes.html', estudiantes=estudiantes)

@app.route('/register_estudiante', methods=['GET', 'POST'])
def register_estudiante():
    if request.method == 'POST':
        cedula = request.form['cedula'].strip()
        nombre = request.form['nombre'].strip()
        sitio = request.form['sitio'].strip()
        nivel_practica = request.form['nivel_practica']
        programa = request.form.get('programa', 'Fisioterapia')
        sede = request.form['sede'].strip()
        correo = request.form.get('correo', '').strip()
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('''
                INSERT INTO estudiantes (cedula, nombre, sitio, nivel_practica, programa, sede, correo)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (cedula, nombre, sitio, nivel_practica, programa, sede, correo))
            conn.commit()
            flash('✅ Estudiante registrado con éxito', 'success')
            return redirect(url_for('estudiantes'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un estudiante con esa cédula', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    return render_template('register_estudiante.html')

@app.route('/edit_estudiante/<int:id>', methods=['GET', 'POST'])
def edit_estudiante(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        cedula = request.form['cedula'].strip()
        nombre = request.form['nombre'].strip()
        sitio = request.form['sitio'].strip()
        nivel_practica = request.form['nivel_practica']
        programa = request.form.get('programa', 'Fisioterapia')
        sede = request.form['sede'].strip()
        correo = request.form.get('correo', '').strip()
        cur.execute('''
            UPDATE estudiantes 
            SET cedula=%s, nombre=%s, sitio=%s, nivel_practica=%s, 
                programa=%s, sede=%s, correo=%s
            WHERE id=%s
        ''', (cedula, nombre, sitio, nivel_practica, programa, sede, correo, id))
        conn.commit()
        flash('✅ Estudiante actualizado', 'success')
        return redirect(url_for('estudiantes'))
    cur.execute("SELECT * FROM estudiantes WHERE id = %s", (id,))
    estudiante = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_estudiante.html', estudiante=estudiante)

@app.route('/delete_estudiante/<int:id>')
def delete_estudiante(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM estudiantes WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Estudiante eliminado', 'danger')
    return redirect(url_for('estudiantes'))


# ==================== DOCENTES CRUD ====================
@app.route('/docentes')
def docentes():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM docentes ORDER BY nombre")
    docentes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('docentes.html', docentes=docentes)

@app.route('/register_docente', methods=['GET', 'POST'])
def register_docente():
    if request.method == 'POST':
        documento = request.form['documento'].strip()
        nombre = request.form['nombre'].strip()
        correo = request.form.get('correo', '').strip()
        conn = get_db_connection()
        cur = conn.cursor()
        try:
            cur.execute('INSERT INTO docentes (documento, nombre, correo) VALUES (%s, %s, %s)',
                        (documento, nombre, correo))
            conn.commit()
            flash('✅ Docente registrado con éxito', 'success')
            return redirect(url_for('docentes'))
        except psycopg2.IntegrityError:
            flash('⚠️ Ya existe un docente con ese documento', 'danger')
        finally:
            cur.close()
            conn.close()
    return render_template('register_docente.html')

@app.route('/edit_docente/<int:id>', methods=['GET', 'POST'])
def edit_docente(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        documento = request.form['documento'].strip()
        nombre = request.form['nombre'].strip()
        correo = request.form.get('correo', '').strip()
        cur.execute('UPDATE docentes SET documento=%s, nombre=%s, correo=%s WHERE id=%s',
                    (documento, nombre, correo, id))
        conn.commit()
        flash('✅ Docente actualizado', 'success')
        return redirect(url_for('docentes'))
    cur.execute("SELECT * FROM docentes WHERE id = %s", (id,))
    docente = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_docente.html', docente=docente)

@app.route('/delete_docente/<int:id>')
def delete_docente(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM docentes WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Docente eliminado', 'danger')
    return redirect(url_for('docentes'))


# ==================== ESCENARIOS CRUD ====================
@app.route('/escenarios')
def escenarios():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM escenarios ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('escenarios.html', escenarios=escenarios)

@app.route('/register_escenario', methods=['GET', 'POST'])
def register_escenario():
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        direccion = request.form.get('direccion', '').strip()
        cupos = int(request.form.get('cupos', 10))
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('INSERT INTO escenarios (nombre, direccion, cupos) VALUES (%s, %s, %s)',
                    (nombre, direccion, cupos))
        conn.commit()
        cur.close()
        conn.close()
        flash('✅ Escenario registrado', 'success')
        return redirect(url_for('escenarios'))
    return render_template('register_escenario.html')

@app.route('/edit_escenario/<int:id>', methods=['GET', 'POST'])
def edit_escenario(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        direccion = request.form.get('direccion', '').strip()
        cupos = int(request.form.get('cupos', 10))
        cur.execute('UPDATE escenarios SET nombre=%s, direccion=%s, cupos=%s WHERE id=%s',
                    (nombre, direccion, cupos, id))
        conn.commit()
        flash('✅ Escenario actualizado', 'success')
        return redirect(url_for('escenarios'))
    cur.execute("SELECT * FROM escenarios WHERE id = %s", (id,))
    escenario = cur.fetchone()
    cur.close()
    conn.close()
    return render_template('edit_escenario.html', escenario=escenario)

@app.route('/delete_escenario/<int:id>')
def delete_escenario(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM escenarios WHERE id = %s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('🗑️ Escenario eliminado', 'danger')
    return redirect(url_for('escenarios'))


# ==================== ASIGNACIONES CRUD ====================
@app.route('/asignaciones')
def asignaciones_list():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT a.id, e.nombre as estudiante, e.cedula, d.nombre as docente,
               es.nombre as escenario, a.rotacion, a.horario, a.fecha_inicio, a.fecha_fin
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY a.fecha_creacion DESC
    ''')
    asignaciones = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('asignaciones.html', asignaciones=asignaciones)

@app.route('/new_assignment', methods=['GET', 'POST'])
def new_assignment():
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        try:
            estudiante_id = int(request.form['estudiante_id'])
            docente_id    = int(request.form['docente_id'])
            escenario_id  = int(request.form['escenario_id'])
            rotacion      = int(request.form['rotacion'])
            horario       = request.form['horario'].strip()
            fecha_inicio  = request.form['fecha_inicio']
            fecha_fin     = request.form['fecha_fin']

            cur.execute('''
                SELECT COUNT(*) AS count FROM asignaciones
                WHERE estudiante_id = %s AND horario = %s
                  AND ((fecha_inicio <= %s AND fecha_fin >= %s)
                    OR (fecha_inicio <= %s AND fecha_fin >= %s))
            ''', (estudiante_id, horario, fecha_fin, fecha_inicio, fecha_inicio, fecha_fin))
            if cur.fetchone()['count'] > 0:
                flash('❌ Conflicto: estudiante ya tiene asignación en ese horario y fechas', 'danger')
                return redirect(url_for('new_assignment'))

            cur.execute('''
                INSERT INTO asignaciones
                    (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (estudiante_id, docente_id, escenario_id, rotacion, horario, fecha_inicio, fecha_fin))
            conn.commit()
            flash('✅ Asignación creada correctamente', 'success')
            return redirect(url_for('asignaciones_list'))
        except Exception as e:
            flash(f'Error al guardar: {str(e)}', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()

    cur.execute("SELECT id, nombre, cedula FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM docentes WHERE estado = 'Activo' ORDER BY nombre")
    docentes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM escenarios WHERE estado = 'Activo' ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('new_assignment.html', estudiantes=estudiantes,
                           docentes=docentes, escenarios=escenarios)

@app.route('/edit_assignment/<int:id>', methods=['GET', 'POST'])
def edit_assignment(id):
    conn = get_db_connection()
    cur = conn.cursor()
    if request.method == 'POST':
        try:
            estudiante_id = int(request.form['estudiante_id'])
            docente_id    = int(request.form['docente_id'])
            escenario_id  = int(request.form['escenario_id'])
            rotacion      = int(request.form['rotacion'])
            horario       = request.form.get('horario', '').strip()
            fecha_inicio  = request.form['fecha_inicio']
            fecha_fin     = request.form['fecha_fin']

            cur.execute('''
                SELECT COUNT(*) AS count FROM asignaciones
                WHERE estudiante_id = %s AND horario = %s AND id <> %s
                  AND ((fecha_inicio <= %s AND fecha_fin >= %s)
                    OR (fecha_inicio <= %s AND fecha_fin >= %s))
            ''', (estudiante_id, horario, id, fecha_fin, fecha_inicio, fecha_inicio, fecha_fin))
            if cur.fetchone()['count'] > 0:
                flash('❌ Conflicto: estudiante ya tiene asignación en ese horario y fechas', 'danger')
                return redirect(url_for('edit_assignment', id=id))

            cur.execute('''
                UPDATE asignaciones
                SET estudiante_id=%s, docente_id=%s, escenario_id=%s,
                    rotacion=%s, horario=%s, fecha_inicio=%s, fecha_fin=%s
                WHERE id=%s
            ''', (estudiante_id, docente_id, escenario_id, rotacion,
                  horario, fecha_inicio, fecha_fin, id))
            conn.commit()
            flash('✅ Asignación actualizada correctamente', 'success')
            return redirect(url_for('asignaciones_list'))
        except Exception as e:
            flash(f'❌ Error al actualizar: {str(e)}', 'danger')
            conn.rollback()
        finally:
            cur.close()
            conn.close()

    cur.execute('SELECT * FROM asignaciones WHERE id = %s', (id,))
    asignacion = cur.fetchone()
    if not asignacion:
        flash('Asignación no encontrada', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('asignaciones_list'))
    cur.execute("SELECT id, nombre, cedula FROM estudiantes ORDER BY nombre")
    estudiantes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM docentes WHERE estado = 'Activo' ORDER BY nombre")
    docentes = cur.fetchall()
    cur.execute("SELECT id, nombre FROM escenarios WHERE estado = 'Activo' ORDER BY nombre")
    escenarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('edit_assignment.html', asignacion=asignacion,
                           estudiantes=estudiantes, docentes=docentes, escenarios=escenarios)

@app.route('/delete_assignment/<int:id>')
def delete_assignment(id):
    conn = cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM asignaciones WHERE id = %s", (id,))
        if not cur.fetchone():
            flash('Asignación no encontrada', 'warning')
            return redirect(url_for('asignaciones_list'))
        cur.execute("DELETE FROM asignaciones WHERE id = %s", (id,))
        conn.commit()
        flash('🗑️ Asignación eliminada correctamente', 'danger')
        return redirect(url_for('asignaciones_list'))
    except Exception as e:
        flash(f'❌ Error al eliminar: {str(e)}', 'danger')
        if conn:
            conn.rollback()
    finally:
        if cur:  cur.close()
        if conn: conn.close()


# ==================== REPORTES ====================

@app.route('/generate_excel_report')
def generate_excel_report():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT e.nombre, e.cedula, es.nombre, d.nombre,
               a.rotacion, a.horario, a.fecha_inicio, a.fecha_fin, es.direccion
        FROM asignaciones a
        JOIN estudiantes e ON a.estudiante_id = e.id
        JOIN docentes d ON a.docente_id = d.id
        JOIN escenarios es ON a.escenario_id = es.id
        ORDER BY e.nombre ASC, a.rotacion ASC
    ''')
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        flash('⚠️ No hay asignaciones para exportar', 'warning')
        return redirect(url_for('index'))

    columns = ["Estudiante", "Documento", "Escenario", "Docente", "Rotación",
               "Horario", "Fecha Inicio", "Fecha Fin", "Dirección"]
    df = pd.DataFrame(rows, columns=columns)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Programación de Prácticas')
        ws = writer.sheets['Programación de Prácticas']
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
        for row in range(2, len(rows) + 2):
            ws.cell(row=row, column=7).number_format = 'DD/MM/YYYY'
            ws.cell(row=row, column=8).number_format = 'DD/MM/YYYY'
    output.seek(0)

    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='Programacion_Practicas.xlsx')


# ──────────────────────────────────────────────────────────────
# PDF DESDE LOS EXCEL ORIGINALES (formato fiel a los horarios)
# ──────────────────────────────────────────────────────────────
@app.route('/generate_pdf_report')
def generate_pdf_report():
    try:
        styles   = getSampleStyleSheet()
        PAGE_W   = landscape(letter)[0] - 2 * cm
        buffer   = io.BytesIO()
        doc      = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                     leftMargin=1*cm, rightMargin=1*cm,
                                     topMargin=1.2*cm, bottomMargin=1.2*cm)
        elements = []

        title_style = ParagraphStyle('main_title', parent=styles['Title'],
                                     fontName='Helvetica-Bold', fontSize=14,
                                     textColor=_AZUL_TITULO, alignment=TA_CENTER,
                                     spaceAfter=4)
        sheet_title_style = ParagraphStyle('sheet_title', parent=styles['Normal'],
                                           fontName='Helvetica-Bold', fontSize=11,
                                           textColor=_BLANCO, alignment=TA_CENTER)
        header_style = ParagraphStyle('header_info', parent=styles['Normal'],
                                      fontName='Helvetica', fontSize=8,
                                      textColor=colors.HexColor('#404040'),
                                      alignment=TA_CENTER)

        elements.append(Paragraph("PROGRAMACIÓN DE PRÁCTICAS ACADÉMICAS 2026-1", title_style))
        elements.append(HRFlowable(width=PAGE_W, thickness=2, color=_AZUL_TITULO))
        elements.append(Spacer(1, 8))

        for fname in EXCEL_FILES:
            fpath = os.path.join(EXCEL_DIR, fname)
            if not os.path.exists(fpath):
                continue                                   # silencia xlsx faltantes

            wb = openpyxl.load_workbook(fpath)
            for ws in wb.worksheets:
                if ws.title.lower() == 'hoja2':
                    continue                               # hoja auxiliar, saltar

                # Banda de título de grupo
                sheet_tbl = Table([[Paragraph(ws.title, sheet_title_style)]],
                                  colWidths=[PAGE_W])
                sheet_tbl.setStyle(TableStyle([
                    ('BACKGROUND',    (0, 0), (-1, -1), _AZUL_TITULO),
                    ('TOPPADDING',    (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(sheet_tbl)
                elements.append(Spacer(1, 4))

                header_lines, rotaciones = _parse_sheet(ws)
                for line in header_lines:
                    if line.strip():
                        elements.append(Paragraph(line, header_style))
                elements.append(Spacer(1, 6))

                for rot in rotaciones:
                    elements += _build_rotation_table(rot, styles)

                elements.append(Spacer(1, 14))
                elements.append(HRFlowable(width=PAGE_W, thickness=1,
                                           color=colors.HexColor('#BFBFBF'), dash=(4, 2)))
                elements.append(Spacer(1, 8))

        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, mimetype='application/pdf',
                         as_attachment=True,
                         download_name='Programacion_Practicas_2026-1.pdf')

    except Exception as e:
        flash(f'Error generando PDF: {str(e)}', 'danger')
        return redirect(url_for('index'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
